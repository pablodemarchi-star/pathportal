import os
import unittest
from io import BytesIO
from types import SimpleNamespace

os.environ["DATABASE_URL"] = "sqlite:///:memory:"

from openpyxl import load_workbook

from app import create_app, db
from app.models import AcademicStaff
from app.routes import (
    INTERN_STAGE_CERTIFICATION_TYPES,
    STAFF_CERTIFICATION_TYPES,
    apply_import_row,
    build_academic_staff_export,
    build_full_annual_certification_export,
    parse_import_workbook,
)


class StaffExportTest(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.drop_all()
        db.create_all()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def test_academic_staff_export_excludes_removed_columns(self):
        member = AcademicStaff(
            id=7,
            status="Active",
            full_name="Jane Staff",
            roles="Examiner",
            email="jane@example.com",
            has_car="Yes",
            full_address_google_maps="742 Evergreen Terrace",
            street_name="Old Street",
            street_number="123",
            city="Springfield",
            postcode="ABC123",
            province="Buenos Aires",
            country="Argentina",
            started_in="2024",
            location_point="https://maps.example.com/location-point",
            cv="https://example.com/cv.pdf",
            profile_picture="https://example.com/profile.jpg",
            dietary_requirements="Vegetarian",
        )

        workbook_bytes = build_academic_staff_export([member], session_counts={member.id: 3})
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook.active

        headers = [cell.value for cell in sheet[3]]
        first_data_row = [cell.value for cell in sheet[4]]

        self.assertNotIn("Location point", headers)
        self.assertNotIn("Updated on", headers)
        self.assertNotIn("Street name", headers)
        self.assertNotIn("Street number", headers)
        self.assertNotIn("Postcode", headers)
        self.assertEqual(headers.index("Full address"), headers.index("Has a car") + 1)
        self.assertEqual(headers.index("Dietary requirements"), headers.index("Profile picture") + 1)
        self.assertEqual(headers.index("Sessions"), headers.index("Started in") + 1)
        self.assertEqual(first_data_row[headers.index("Full address")], member.full_address_google_maps)
        self.assertEqual(first_data_row[headers.index("Dietary requirements")], member.dietary_requirements)
        self.assertEqual(first_data_row[headers.index("Sessions")], 3)
        self.assertNotIn("https://maps.example.com/location-point", first_data_row)
        self.assertNotIn("Old Street", first_data_row)
        self.assertNotIn("123", first_data_row)
        self.assertNotIn("ABC123", first_data_row)
        self.assertEqual(sheet.cell(row=4, column=headers.index("CV") + 1).hyperlink.target, member.cv)
        self.assertEqual(
            sheet.cell(row=4, column=headers.index("Profile picture") + 1).hyperlink.target,
            member.profile_picture,
        )

    def test_exported_full_address_is_imported(self):
        source_member = AcademicStaff(
            id=8,
            status="Active",
            title="Prof.",
            full_name="John Staff",
            roles="Supervisor",
            phone="555-0101",
            email="john@example.com",
            has_car="No",
            full_address_google_maps="Av. Siempre Viva 123",
            city="CABA",
            province="Buenos Aires",
            country="Argentina",
            started_in="2025",
            cv="https://example.com/john-cv.pdf",
            account_id="ACC-8",
            account_owner="Path",
            profile_picture="https://example.com/john-profile.jpg",
            dietary_requirements="Gluten-free meal",
        )
        workbook_bytes = build_academic_staff_export([source_member], session_counts={source_member.id: 2})

        payload = parse_import_workbook(BytesIO(workbook_bytes), update_empty_fields=False)
        self.assertEqual(payload["unknown_headers"], [])
        self.assertEqual(payload["summary"]["ready"], 1)

        row_data = payload["rows"][0]["data"]
        self.assertEqual(row_data["Full address"], source_member.full_address_google_maps)
        self.assertEqual(row_data["Dietary requirements"], source_member.dietary_requirements)

        imported_member = AcademicStaff()
        apply_import_row(imported_member, row_data, update_empty_fields=True)
        self.assertEqual(imported_member.full_address_google_maps, source_member.full_address_google_maps)
        self.assertEqual(imported_member.dietary_requirements, source_member.dietary_requirements)

    def test_import_requires_complete_member_fields_except_seniority_and_history(self):
        source_member = AcademicStaff(
            id=9,
            status="Active",
            full_name="Incomplete Staff",
            roles="Examiner",
            email="incomplete@example.com",
            has_car="Yes",
            city="CABA",
            province="Buenos Aires",
            country="Argentina",
        )
        workbook_bytes = build_academic_staff_export([source_member], session_counts={source_member.id: 0})

        payload = parse_import_workbook(BytesIO(workbook_bytes), update_empty_fields=False)

        self.assertEqual(payload["summary"]["ready"], 0)
        self.assertEqual(payload["summary"]["errors"], 1)
        row = payload["rows"][0]
        self.assertEqual(row["action"], "Error")
        self.assertIn("Title is required.", row["errors"])
        self.assertIn("Phone is required.", row["errors"])
        self.assertIn("Full address is required.", row["errors"])
        self.assertIn("CV is required.", row["errors"])
        self.assertIn("Account ID is required.", row["errors"])
        self.assertIn("Account owner is required.", row["errors"])
        self.assertIn("Profile picture is required.", row["errors"])
        self.assertIn("Started in is required.", row["errors"])
        self.assertNotIn("Seniority is required.", row["errors"])
        self.assertNotIn("History is required.", row["errors"])

    def test_certification_export_uses_full_staff_export_design(self):
        member = AcademicStaff(
            id=10,
            status="Active",
            title="Dr.",
            full_name="Examiner Staff",
            roles="Examiner",
            phone="555-0110",
            email="examiner@example.com",
            has_car="Yes",
            full_address_google_maps="Main Road 100",
            city="Moreno",
            province="Buenos Aires",
            country="Argentina",
            cv="https://example.com/examiner-cv.pdf",
            interview="First note",
            account_id="ACC-10",
            account_owner="Path",
            profile_picture="https://example.com/examiner.jpg",
            started_in="2026",
            seniority=True,
        )
        workbook_bytes = build_full_annual_certification_export(
            [member],
            {},
            export_title="Examiner Certification Export",
            sheet_title="Examiner Certification",
            status_header="Certification status",
            certification_types=STAFF_CERTIFICATION_TYPES,
            session_counts={member.id: 4},
            remote_training_selections={member.id: SimpleNamespace(status="Certified")},
            annual_meeting_selections={member.id: SimpleNamespace(status="Attended")},
            fut_selections={member.id: [SimpleNamespace(option_name="FUT 1", status="completed")]},
            fut2_selections={},
        )

        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[3]]
        first_data_row = [cell.value for cell in sheet[4]]

        self.assertEqual(sheet.title, "Examiner Certification")
        self.assertEqual(sheet["A1"].value, "Examiner Certification Export")
        self.assertEqual(headers[:20], [
            "Status",
            "Title",
            "Full name",
            "Roles",
            "Phone",
            "Email",
            "Has a car",
            "Full address",
            "City",
            "Province",
            "Country",
            "CV",
            "History",
            "Account ID",
            "Account owner",
            "Profile picture",
            "Dietary requirements",
            "Started in",
            "Sessions",
            "Seniority",
        ])
        self.assertIn("Certification status", headers)
        self.assertIn("Annual meeting", headers)
        self.assertIn("Remote training", headers)
        self.assertIn("FUT", headers)
        self.assertEqual(first_data_row[headers.index("Full address")], member.full_address_google_maps)
        self.assertEqual(first_data_row[headers.index("Sessions")], 4)
        self.assertEqual(first_data_row[headers.index("Certification status")], "Certified")
        self.assertEqual(first_data_row[headers.index("Annual meeting")], "Attended")
        self.assertEqual(first_data_row[headers.index("Remote training")], "Certified")
        self.assertEqual(first_data_row[headers.index("FUT")], "FUT 1 (completed)")
        self.assertEqual(sheet.cell(row=4, column=headers.index("CV") + 1).hyperlink.target, member.cv)
        self.assertEqual(
            sheet.cell(row=4, column=headers.index("Profile picture") + 1).hyperlink.target,
            member.profile_picture,
        )

    def test_intern_stages_export_adds_stage_columns_to_full_staff_design(self):
        member = AcademicStaff(
            id=11,
            status="Active",
            full_name="Intern Staff",
            roles="Intern",
            email="intern@example.com",
            has_car="No",
            full_address_google_maps="Training Road 200",
            started_in="2026",
        )
        workbook_bytes = build_full_annual_certification_export(
            [member],
            {},
            export_title="Intern Stages Export",
            sheet_title="Intern Stages",
            status_header="Stage status",
            certification_types=INTERN_STAGE_CERTIFICATION_TYPES,
            session_counts={member.id: 1},
            remote_training_selections={member.id: SimpleNamespace(status="Completed")},
            stage_3_selections={member.id: SimpleNamespace(status="In progress")},
            stage_2_selections={},
            fut_selections={},
            fut2_selections={},
        )

        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[3]]
        first_data_row = [cell.value for cell in sheet[4]]

        self.assertEqual(headers[:20], [
            "Status",
            "Title",
            "Full name",
            "Roles",
            "Phone",
            "Email",
            "Has a car",
            "Full address",
            "City",
            "Province",
            "Country",
            "CV",
            "History",
            "Account ID",
            "Account owner",
            "Profile picture",
            "Dietary requirements",
            "Started in",
            "Sessions",
            "Seniority",
        ])
        self.assertIn("Stage status", headers)
        self.assertIn("Stage 1", headers)
        self.assertIn("Stage 2", headers)
        self.assertIn("FUT", headers)
        self.assertIn("Stage 3", headers)
        self.assertEqual(first_data_row[headers.index("Stage status")], "In progress")
        self.assertEqual(first_data_row[headers.index("Stage 1")], "Completed")
        self.assertEqual(first_data_row[headers.index("Stage 2")], "In progress")
        self.assertEqual(first_data_row[headers.index("FUT")], "*")
        self.assertEqual(first_data_row[headers.index("Stage 3")], "*")


if __name__ == "__main__":
    unittest.main()
